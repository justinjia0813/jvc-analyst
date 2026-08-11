#!/usr/bin/env python3
"""Validate jvc skill route and output eval fixtures.

This is deterministic fixture validation. It checks that eval cases point to
real skills, prompts carry their intended route signals, and output assertions
still match tracked templates/examples. It does not claim model-executed eval
evidence.
"""

from __future__ import annotations

import json
import runpy
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
MARKET_VALIDATOR = ROOT / "skills" / "jvc-market-sizing" / "scripts" / "validate_csv.py"
RESEARCHCTL = ROOT / "skills" / "jvc-research-core" / "scripts" / "researchctl.py"
MARKET_EVAL_PATHS = {
    "baseline": "evals/research-core/baselines/industrial_vision_software_market_20260729.csv",
    "candidate": "evals/research-core/candidates/industrial_vision_software_market_20260729.csv",
    "run_artifact": "evals/research-core/runs/market-model-fact-vs-assumption/industrial_vision_software_market_20260729.csv",
    "evidence_registry": "evals/research-core/runs/market-model-fact-vs-assumption/evidence_registry.jsonl",
    "audit_json": "evals/research-core/runs/market-model-fact-vs-assumption/audit.json",
}


def load_json(relative_path: str) -> dict[str, Any]:
    path = ROOT / relative_path
    if not path.is_file():
        raise AssertionError(f"missing eval file: {relative_path}")
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def read_text(relative_path: str) -> str:
    path = ROOT / relative_path
    if not path.is_file():
        raise AssertionError(f"missing file: {relative_path}")
    return path.read_text(encoding="utf-8")


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def skill_path(skill: str) -> Path:
    return ROOT / "skills" / skill / "SKILL.md"


def require_skill(skill: str) -> str:
    path = skill_path(skill)
    require(path.is_file(), f"missing skill: {skill}")
    text = path.read_text(encoding="utf-8")
    require(f"name: {skill}" in text, f"{skill} SKILL.md missing matching name")
    return text


def require_unique_ids(cases: list[dict[str, Any]], source: str) -> None:
    seen: set[str] = set()
    for case in cases:
        case_id = case.get("id")
        require(isinstance(case_id, str) and case_id, f"{source} case missing id")
        require(case_id not in seen, f"duplicate {source} case id: {case_id}")
        seen.add(case_id)


def canonical_assertion_paths(case: dict[str, Any]) -> set[str]:
    case_id = case["id"]
    assertions = case.get("assertions")
    require(isinstance(assertions, list) and assertions, f"{case_id}: missing assertions")

    paths: set[str] = set()
    root = ROOT.resolve()
    for index, assertion in enumerate(assertions):
        require(isinstance(assertion, dict), f"{case_id}: assertion {index} must be an object")
        relative_path = assertion.get("path")
        require(
            isinstance(relative_path, str) and relative_path,
            f"{case_id}: assertion {index} missing path",
        )
        require(not Path(relative_path).is_absolute(), f"{case_id}: assertion path must be relative: {relative_path}")
        resolved = (ROOT / relative_path).resolve()
        try:
            canonical = resolved.relative_to(root).as_posix()
        except ValueError as exc:
            raise AssertionError(f"{case_id}: assertion path escapes repository: {relative_path}") from exc
        require(
            relative_path == canonical,
            f"{case_id}: assertion path must be canonical: {relative_path!r} != {canonical!r}",
        )
        paths.add(canonical)
    return paths


def check_prescreen_case_independence(
    cases: list[dict[str, Any]], paths_by_case: dict[str, set[str]]
) -> None:
    expected_artifacts = {
        "prescreen-supported-markdown-contract": "examples/prescreen-example.md",
        "prescreen-missing-data-markdown-contract": "examples/prescreen-missing-data-example.md",
    }
    prescreen_cases = {case["id"]: case for case in cases if case.get("skill") == "jvc-prescreen"}
    required_case_ids = set(expected_artifacts)
    require(
        required_case_ids <= set(prescreen_cases),
        f"missing required Pre-Screen output cases: {sorted(required_case_ids - set(prescreen_cases))}",
    )

    for case_id, artifact_path in expected_artifacts.items():
        paths = paths_by_case[case_id]
        require(artifact_path in paths, f"{case_id}: missing artifact {artifact_path}")
        other_artifacts = {
            path for other_id, path in expected_artifacts.items() if other_id != case_id
        }
        require(
            paths.isdisjoint(other_artifacts),
            f"{case_id}: must not include other Pre-Screen artifacts {sorted(paths & other_artifacts)}",
        )


def load_jsonl(relative_path: str) -> list[dict[str, Any]]:
    path = ROOT / relative_path
    require(path.is_file(), f"missing eval file: {relative_path}")
    records: list[dict[str, Any]] = []
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            record = json.loads(line)
        except json.JSONDecodeError as exc:
            raise AssertionError(f"{relative_path}:{line_number}: invalid JSON") from exc
        require(isinstance(record, dict), f"{relative_path}:{line_number}: expected object")
        records.append(record)
    return records


def run_market_validator(path: Path) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        (sys.executable, str(MARKET_VALIDATOR), str(path)),
        cwd=ROOT,
        capture_output=True,
        check=False,
        text=True,
    )


def path_ends_with(path_value: object, relative_path: str) -> bool:
    if not isinstance(path_value, str):
        return False
    actual = Path(path_value).parts
    expected = Path(relative_path).parts
    return len(actual) >= len(expected) and actual[-len(expected) :] == expected


def audit_binding_matches(
    entry: dict[str, Any],
    recomputed: dict[str, Any],
    records: list[dict[str, Any]],
    research_core: dict[str, Any],
) -> bool:
    if not research_core["_audit_shape_is_valid"](entry):
        return False
    if entry["audit_key"] != research_core["audit_key"](entry):
        return False
    exact_fields = (
        "schema_version",
        "skill",
        "ledger_sequence",
        "ledger_prefix_fingerprint",
        "dependency_audits",
        "profile_fingerprint",
        "core_runtime_fingerprint",
        "status",
        "findings",
    )
    if any(entry[field] != recomputed[field] for field in exact_fields):
        return False
    if len(entry["artifacts"]) != 1 or len(recomputed["artifacts"]) != 1:
        return False
    if entry["artifacts"][0]["fingerprint"] != recomputed["artifacts"][0]["fingerprint"]:
        return False
    if not path_ends_with(entry["artifacts"][0]["path"], MARKET_EVAL_PATHS["run_artifact"]):
        return False
    if not path_ends_with(entry["audit_path"], MARKET_EVAL_PATHS["audit_json"]):
        return False

    portable = json.loads(json.dumps(entry, ensure_ascii=False))
    portable["audit_path"] = str((ROOT / MARKET_EVAL_PATHS["audit_json"]).resolve())
    portable["artifacts"][0]["path"] = str(
        (ROOT / MARKET_EVAL_PATHS["run_artifact"]).resolve()
    )
    portable["audit_key"] = research_core["audit_key"](portable)
    return research_core["saved_audit_is_valid"](portable, records)


def recompute_market_audit(research_core: dict[str, Any]) -> dict[str, Any]:
    run_dir = ROOT / "evals" / "research-core" / "runs" / "market-model-fact-vs-assumption"
    run_artifact = ROOT / MARKET_EVAL_PATHS["run_artifact"]
    with tempfile.TemporaryDirectory(prefix="jvc-market-audit-") as directory:
        temporary_run = Path(directory)
        shutil.copy2(run_dir / "evidence_registry.jsonl", temporary_run / "evidence_registry.jsonl")
        temporary_artifact = temporary_run / run_artifact.name
        shutil.copy2(run_artifact, temporary_artifact)
        return research_core["audit_run"](
            temporary_run,
            "jvc-market-sizing",
            [temporary_artifact],
        )


def check_invalid_market_csv_regression(run_artifact: str) -> None:
    with tempfile.TemporaryDirectory(prefix="jvc-invalid-market-csv-") as directory:
        invalid = Path(directory) / "market-sizing.csv"
        rows = [
            line
            for line in run_artifact.splitlines()
            if not line.startswith("orthogonality_check,ORTHO_1,")
        ]
        require(len(rows) + 1 == len(run_artifact.splitlines()), "invalid CSV regression fixture did not remove ORTHO_1")
        invalid.write_text("\n".join(rows) + "\n", encoding="utf-8")
        result = run_market_validator(invalid)
        require(
            result.returncode != 0,
            "Market Sizing validator accepted CSV without orthogonality_check",
        )


def check_p1_eval_integration(
    cases: list[dict[str, Any]], paths_by_case: dict[str, set[str]]
) -> None:
    by_id = {case["id"]: case for case in cases}
    for legacy_id in (
        "market-sizing-workbook-contract",
        "comps-dd-workbook-contract",
        "roi-modeler-workbook-contract",
    ):
        require(legacy_id not in by_id, f"obsolete {legacy_id} output case is active")
    market_case = by_id.get("market-sizing-csv-contract")
    require(isinstance(market_case, dict), "missing Market Sizing CSV output case")
    require(market_case.get("artifact_family") == "csv", "Market Sizing output case must use CSV")
    require(
        {
            "templates/market-sizing-template.csv",
            "examples/market-sizing-example.csv",
            "skills/jvc-market-sizing/scripts/validate_csv.py",
            "skills/jvc-market-sizing/scripts/check_package.py",
        }
        <= paths_by_case["market-sizing-csv-contract"],
        "Market Sizing output case missing active CSV assets",
    )

    knowledge_paths = paths_by_case.get("knowledge-tree-builder-artifact-contract", set())
    require(
        {
            "skills/jvc-knowledge-tree-builder/scripts/validate_output.py",
            "skills/jvc-knowledge-tree-builder/scripts/check_package.py",
            "examples/knowledge-tree-example/knowledge_tree.md",
            "examples/knowledge-tree-example/knowledge_graph.mmd",
            "examples/knowledge-tree-example/nodes.json",
            "examples/knowledge-tree-example/evidence_index.md",
            "examples/knowledge-tree-example/open_questions.md",
        }
        <= knowledge_paths,
        "Knowledge Tree output case missing validator or fixed five-file example",
    )

    research_cases = load_json("evals/research-core/cases.json").get("cases", [])
    indexed = {case.get("id"): case for case in research_cases if isinstance(case, dict)}
    market_index = indexed.get("market-model-fact-vs-assumption")
    require(isinstance(market_index, dict), "missing Research Core market model case")
    for key, relative_path in MARKET_EVAL_PATHS.items():
        require(market_index.get(key) == relative_path, f"Research Core market case {key} mismatch")
        require((ROOT / relative_path).is_file(), f"missing active Research Core CSV asset: {relative_path}")

    for key in ("baseline", "candidate", "run_artifact"):
        relative_path = MARKET_EVAL_PATHS[key]
        result = run_market_validator(ROOT / relative_path)
        require(
            result.returncode == 0,
            f"invalid active Market Sizing CSV {relative_path}: {result.stdout}{result.stderr}",
        )

    output_cases = load_jsonl("evals/research-core/output_cases.jsonl")
    market_outputs = [case for case in output_cases if case.get("id") == "industrial-vision-market-model-output"]
    require(len(market_outputs) == 1, "Research Core market output case must be unique")
    output_case = market_outputs[0]
    metadata = output_case.get("metadata", {})
    require(metadata.get("artifact_type") == "csv", "Research Core market output metadata must use CSV")
    for key, relative_path in MARKET_EVAL_PATHS.items():
        require(metadata.get(key) == relative_path, f"Research Core market output {key} mismatch")
    baseline = (ROOT / MARKET_EVAL_PATHS["baseline"]).read_text(encoding="utf-8")
    candidate = (ROOT / MARKET_EVAL_PATHS["candidate"]).read_text(encoding="utf-8")
    run_artifact = (ROOT / MARKET_EVAL_PATHS["run_artifact"]).read_text(encoding="utf-8")
    require(output_case.get("baseline_output") == baseline, "Research Core baseline output does not match CSV fixture")
    require(output_case.get("with_skill_output") == candidate, "Research Core candidate output does not match CSV fixture")
    require(candidate == run_artifact, "Research Core candidate and audited run artifact differ")
    check_invalid_market_csv_regression(run_artifact)
    active_case_text = json.dumps(market_index, ensure_ascii=False) + json.dumps(output_case, ensure_ascii=False)
    require(".xlsx" not in active_case_text and "cell-dump" not in active_case_text, "active Research Core case references a historical workbook fixture")

    research_core = runpy.run_path(str(RESEARCHCTL))
    run_dir = ROOT / "evals" / "research-core" / "runs" / "market-model-fact-vs-assumption"
    records = research_core["load_registry"](run_dir)
    recomputed = recompute_market_audit(research_core)
    audit = load_json(MARKET_EVAL_PATHS["audit_json"])
    audits = audit.get("audits", [])
    require(len(audits) == 1, "Research Core market audit must be unique")
    entry = audits[0]
    require(
        audit_binding_matches(entry, recomputed, records, research_core),
        "Research Core market audit binding mismatch",
    )
    tampered = json.loads(json.dumps(entry, ensure_ascii=False))
    tampered["ledger_prefix_fingerprint"] = "0" * 64
    tampered["audit_key"] = research_core["audit_key"](tampered)
    require(
        not audit_binding_matches(tampered, recomputed, records, research_core),
        "Research Core audit-binding tamper regression was accepted",
    )
    require(
        any(
            record.get("record_type") == "source"
            and record.get("source_class") == "user-document"
            and record.get("location") == "input/scope.json"
            for record in records
        ),
        "Research Core market ledger missing the user-assumption evidence pointer",
    )


def require_signal_list(case: dict[str, Any], key: str, *, required: bool = False) -> list[str]:
    case_id = case["id"]
    signals = case.get(key, [])
    require(isinstance(signals, list), f"{case_id}: {key} must be a list")
    require(not required or signals, f"{case_id}: {key} must be non-empty")
    require(
        all(isinstance(signal, str) and signal.strip() for signal in signals),
        f"{case_id}: {key} must contain only non-empty strings",
    )
    require(len(signals) == len(set(signals)), f"{case_id}: {key} contains duplicate signals")
    return signals


def all_skill_names() -> set[str]:
    return {path.parent.name for path in (ROOT / "skills").glob("jvc-*/SKILL.md")}


def user_invocable_skill_names() -> set[str]:
    names: set[str] = set()
    for path in (ROOT / "skills").glob("jvc-*/SKILL.md"):
        text = path.read_text(encoding="utf-8")
        parts = text.split("---", 2)
        require(len(parts) == 3 and not parts[0].strip(), f"{path.parent.name}: malformed frontmatter")
        frontmatter_lines = {line.strip() for line in parts[1].splitlines()}
        if "user_invocable: false" not in frontmatter_lines:
            names.add(path.parent.name)
    return names


def check_trigger_cases() -> int:
    data = load_json("evals/trigger_cases.json")
    cases = data.get("cases")
    require(data.get("schema_version") == 1, "trigger cases schema_version must be 1")
    require(isinstance(cases, list) and cases, "trigger cases must be a non-empty list")
    require_unique_ids(cases, "trigger")

    expected_pairs = {
        ("jvc-deal-flow", "jvc-prescreen"),
        ("jvc-deal-flow", "jvc-ic-memo"),
        ("jvc-prescreen", "jvc-deal-flow"),
        ("jvc-prescreen", "jvc-ic-memo"),
        ("jvc-talk-notes", "jvc-meeting-notes"),
        ("jvc-meeting-notes", "jvc-talk-notes"),
        ("jvc-bull-case", "jvc-ic-memo"),
        ("jvc-bear-case", "jvc-bull-case"),
        ("jvc-ic-memo", "jvc-bull-case"),
        ("jvc-track-research", "jvc-comps-dd"),
        ("jvc-track-research", "jvc-research-report"),
        ("jvc-research-report", "jvc-track-research"),
        ("jvc-research-report", "jvc-ic-memo"),
        ("jvc-comps-dd", "jvc-track-research"),
        ("jvc-market-sizing", "jvc-track-research"),
        ("jvc-roi-modeler", "jvc-market-sizing"),
        ("jvc-invoice-manager", "jvc-comps-dd"),
    }
    covered_pairs: set[tuple[str, str]] = set()
    no_route_count = 0

    for case in cases:
        case_id = case["id"]
        prompt = case.get("prompt")
        require(isinstance(prompt, str) and prompt.strip(), f"{case_id}: missing prompt")
        for signal in case.get("prompt_signals", []):
            require(signal in prompt, f"{case_id}: prompt missing signal {signal!r}")

        expected_skill = case.get("expected_skill")
        if expected_skill is None:
            no_route_count += 1
            reason = case.get("no_route_reason")
            require(isinstance(reason, str) and reason.strip(), f"{case_id}: missing no_route_reason")
            for skill in case.get("should_not_trigger", []):
                require_skill(skill)
            continue

        require(isinstance(expected_skill, str) and expected_skill, f"{case_id}: missing expected_skill")
        skill_text = require_skill(expected_skill)
        for signal in require_signal_list(case, "skill_contract_signals", required=True):
            require(signal in skill_text, f"{case_id}: {expected_skill} missing contract signal {signal!r}")
        for signal in require_signal_list(case, "skill_contract_forbidden_signals"):
            require(signal not in skill_text, f"{case_id}: {expected_skill} contains forbidden contract signal {signal!r}")

        neighbors = case.get("near_neighbors", [])
        require(isinstance(neighbors, list) and neighbors, f"{case_id}: expected at least one near neighbor")
        for neighbor in neighbors:
            neighbor_skill = neighbor.get("skill")
            why_not = neighbor.get("why_not")
            require(isinstance(neighbor_skill, str) and neighbor_skill, f"{case_id}: neighbor missing skill")
            require_skill(neighbor_skill)
            require(isinstance(why_not, str) and why_not.strip(), f"{case_id}: neighbor {neighbor_skill} missing why_not")
            covered_pairs.add((expected_skill, neighbor_skill))

    missing_pairs = expected_pairs - covered_pairs
    require(not missing_pairs, f"missing near-neighbor trigger coverage: {sorted(missing_pairs)}")
    require(no_route_count >= 1, "trigger evals should include at least one no-route teaching/explanation case")
    routed_skills = {case.get("expected_skill") for case in cases if case.get("expected_skill")}
    missing_skills = user_invocable_skill_names() - routed_skills
    require(not missing_skills, f"missing trigger coverage for skills: {sorted(missing_skills)}")
    return len(cases)


def check_assertion(case_id: str, assertion: dict[str, Any]) -> None:
    assertion_type = assertion.get("type")
    relative_path = assertion.get("path")
    if assertion_type in {"file_exists", "file_tracked", "contains", "contains_any", "not_contains_any", "workbook_sheets"}:
        require(isinstance(relative_path, str) and relative_path, f"{case_id}: assertion missing path")

    if assertion_type == "file_exists":
        require((ROOT / relative_path).is_file(), f"{case_id}: missing file {relative_path}")
        return

    if assertion_type == "file_tracked":
        require((ROOT / relative_path).is_file(), f"{case_id}: missing file {relative_path}")
        tracked = subprocess.run(
            ("git", "ls-files", "--error-unmatch", "--", relative_path),
            cwd=ROOT,
            capture_output=True,
            check=False,
        )
        require(tracked.returncode == 0, f"{case_id}: untracked file {relative_path}")
        return

    if assertion_type == "contains":
        text = assertion.get("text")
        require(isinstance(text, str) and text, f"{case_id}: contains assertion missing text")
        require(text in read_text(relative_path), f"{case_id}: {relative_path} missing {text!r}")
        return

    if assertion_type == "contains_any":
        texts = assertion.get("texts")
        require(isinstance(texts, list) and texts, f"{case_id}: contains_any missing texts")
        haystack = read_text(relative_path)
        require(any(isinstance(text, str) and text in haystack for text in texts), f"{case_id}: {relative_path} missing any of {texts!r}")
        return

    if assertion_type == "not_contains_any":
        texts = assertion.get("texts")
        require(isinstance(texts, list), f"{case_id}: not_contains_any missing texts")
        haystack = read_text(relative_path)
        found = [text for text in texts if isinstance(text, str) and text in haystack]
        require(not found, f"{case_id}: {relative_path} contains forbidden text {found!r}")
        return

    if assertion_type == "workbook_sheets":
        sheets = assertion.get("sheets")
        require(isinstance(sheets, list) and sheets, f"{case_id}: workbook_sheets missing sheets")
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise AssertionError("openpyxl is required for workbook_sheets assertions") from exc
        workbook_path = ROOT / relative_path
        require(workbook_path.is_file(), f"{case_id}: missing workbook {relative_path}")
        workbook = load_workbook(workbook_path, read_only=True)
        actual = set(workbook.sheetnames)
        missing = [sheet for sheet in sheets if sheet not in actual]
        require(not missing, f"{case_id}: {relative_path} missing workbook sheets {missing!r}")
        return

    raise AssertionError(f"{case_id}: unknown assertion type {assertion_type!r}")


def check_output_cases() -> int:
    data = load_json("evals/output/cases.json")
    cases = data.get("cases")
    require(data.get("schema_version") == 1, "output cases schema_version must be 1")
    require(isinstance(cases, list) and cases, "output cases must be a non-empty list")
    require_unique_ids(cases, "output")

    paths_by_case = {case["id"]: canonical_assertion_paths(case) for case in cases}
    check_prescreen_case_independence(cases, paths_by_case)
    check_p1_eval_integration(cases, paths_by_case)

    required_families = {"markdown", "csv", "docx", "excel_pdf_archive", "research_pdf"}
    families: set[str] = set()
    for case in cases:
        case_id = case["id"]
        skill = case.get("skill")
        artifact_family = case.get("artifact_family")
        assertions = case["assertions"]
        require(isinstance(skill, str) and skill, f"{case_id}: missing skill")
        require_skill(skill)
        require(isinstance(artifact_family, str) and artifact_family, f"{case_id}: missing artifact_family")
        families.add(artifact_family)
        for assertion in assertions:
            check_assertion(case_id, assertion)

    missing_families = required_families - families
    require(not missing_families, f"missing output artifact families: {sorted(missing_families)}")
    output_skills = {case.get("skill") for case in cases}
    missing_skills = all_skill_names() - output_skills
    require(not missing_skills, f"missing output coverage for skills: {sorted(missing_skills)}")
    return len(cases)


def main() -> int:
    try:
        trigger_count = check_trigger_cases()
        output_count = check_output_cases()
    except AssertionError as exc:
        print(f"skill eval check failed: {exc}", file=sys.stderr)
        return 1

    print(f"skill eval fixtures passed: {trigger_count} trigger cases, {output_count} output cases")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
