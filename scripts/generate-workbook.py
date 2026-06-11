#!/usr/bin/env python3
"""Generate an XLSX workbook from a skill Markdown sheet schema."""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - exercised by environment setup
    raise SystemExit(
        "openpyxl is required. Install it with: python3 -m pip install openpyxl"
    ) from exc


SHEET_RE = re.compile(r"^## Sheet:\s*(.+?)\s*$")


def split_markdown_row(line: str) -> list[str]:
    cells = line.strip().strip("|").split("|")
    return [cell.strip() for cell in cells]


def is_separator_row(cells: list[str]) -> bool:
    return bool(cells) and all(re.fullmatch(r":?-{3,}:?", cell) for cell in cells)


def parse_template(template_path: Path) -> dict[str, list[str]]:
    sheets: dict[str, list[str]] = {}
    current_sheet: str | None = None

    for raw_line in template_path.read_text(encoding="utf-8").splitlines():
        sheet_match = SHEET_RE.match(raw_line)
        if sheet_match:
            current_sheet = sheet_match.group(1)
            sheets[current_sheet] = []
            continue

        if not current_sheet or not raw_line.strip().startswith("|"):
            continue

        cells = split_markdown_row(raw_line)
        if not cells or cells[0] == "字段" or is_separator_row(cells):
            continue

        field_name = cells[0]
        if field_name:
            sheets[current_sheet].append(field_name)

    empty_sheets = [name for name, fields in sheets.items() if not fields]
    if empty_sheets:
        joined = ", ".join(empty_sheets)
        raise ValueError(f"template sheets without fields: {joined}")
    if not sheets:
        raise ValueError("template contains no '## Sheet:' sections")

    return sheets


def size_columns(ws) -> None:
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def build_workbook(sheets: dict[str, list[str]], output_path: Path) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for sheet_name, headers in sheets.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        size_columns(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate an XLSX workbook from a Markdown sheet schema."
    )
    parser.add_argument("template", type=Path, help="Markdown template path")
    parser.add_argument("output", type=Path, help="Output .xlsx path")
    args = parser.parse_args()

    if args.output.suffix.lower() != ".xlsx":
        parser.error("output path must end with .xlsx")

    try:
        sheets = parse_template(args.template)
        build_workbook(sheets, args.output)
    except Exception as exc:
        print(f"failed to generate workbook: {exc}", file=sys.stderr)
        return 1

    print(f"generated {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
