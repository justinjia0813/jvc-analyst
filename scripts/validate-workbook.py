#!/usr/bin/env python3
"""Validate that an XLSX workbook matches a skill Markdown sheet schema."""

from __future__ import annotations

import argparse
import importlib.util
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - exercised by environment setup
    raise SystemExit(
        "openpyxl is required. Install it with: python3 -m pip install openpyxl"
    ) from exc

def load_parse_template():
    script_path = Path(__file__).with_name("generate-workbook.py")
    spec = importlib.util.spec_from_file_location("generate_workbook", script_path)
    if not spec or not spec.loader:
        raise RuntimeError(f"cannot load parser from {script_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module.parse_template


def read_headers(workbook_path: Path) -> dict[str, list[str]]:
    wb = load_workbook(workbook_path, data_only=False)
    headers: dict[str, list[str]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers[sheet_name] = [str(value) for value in first_row if value is not None]
    return headers


def validate(workbook_path: Path, template_path: Path) -> list[str]:
    parse_template = load_parse_template()
    expected = parse_template(template_path)
    actual = read_headers(workbook_path)
    errors: list[str] = []

    missing_sheets = [name for name in expected if name not in actual]
    extra_sheets = [name for name in actual if name not in expected]

    for sheet_name in missing_sheets:
        errors.append(f"missing sheet: {sheet_name}")
    for sheet_name in extra_sheets:
        errors.append(f"unexpected sheet: {sheet_name}")

    for sheet_name, expected_headers in expected.items():
        if sheet_name not in actual:
            continue
        actual_headers = actual[sheet_name]
        if actual_headers != expected_headers:
            errors.append(
                f"header mismatch in {sheet_name}: "
                f"expected {expected_headers}, got {actual_headers}"
            )

    return errors


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Validate an XLSX workbook against a Markdown sheet schema."
    )
    parser.add_argument("workbook", type=Path, help="Workbook .xlsx path")
    parser.add_argument("template", type=Path, help="Markdown template path")
    args = parser.parse_args()

    if not args.workbook.exists():
        parser.error(f"workbook does not exist: {args.workbook}")
    if not args.template.exists():
        parser.error(f"template does not exist: {args.template}")

    errors = validate(args.workbook, args.template)
    if errors:
        for error in errors:
            print(error, file=sys.stderr)
        return 1

    print(f"validated {args.workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
